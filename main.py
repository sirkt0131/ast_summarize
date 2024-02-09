import concurrent.futures
from data.model_info import ModelInfo
from data.selection import GetList
from src.chat_model import ChatModel
from src.utils import (clean_text, split_text, sum_cost, RateLimiter)
import streamlit as st
from openpyxl import load_workbook
from src.summarize import summarize
from src.logging import ThreadSafeLogger

model = st.selectbox('select model', GetList.get_model_list())
model_info = ModelInfo.get_model_info(model)
max_word = int(model_info["max_token"]/3)
chat_model = ChatModel(model_info["model_name"])
logger = ThreadSafeLogger()
# Calculate the delay based on your rate limit
rate_limit_per_minute = 20
delay = 60.0 / rate_limit_per_minute

uploaded_file = st.file_uploader("Astrategy事例集をアップロードしてください")

if uploaded_file is not None:
    book = load_workbook(uploaded_file)
    sheet = book[book.sheetnames[0]]

    coord = [] # (col, row, value, link)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.hyperlink:  # セルにハイパーリンクがあるかチェック
                coord.append([cell.column, cell.row, cell.hyperlink, cell.value])
        if len(coord) >= 20: # Max20行
            break
    
    summarize_btn = st.button('summarize')

    if summarize_btn:
        my_bar = st.progress(0)
        costs = []
        # for index, c in enumerate(coord):
        #     url = c[2].target
        #     col = c[0]
        #     row = c[1]
        #     title = c[3]

        #     cell = sheet.cell(column=col+2, row=row)
        #     # summarize by ChatGPT
        #     result, cost = summarize(url, max_word, chat_model, model_info)
        #     cell.value = result
        #     st.write(f'[{title}]({url})')
        #     st.write(result)
        #     costs.append(cost)
        #     my_bar.progress(int((index+1)/len(coord)*100))
        
        # https://github.com/openai/openai-cookbook/blob/main/examples/How_to_handle_rate_limits.ipynb
        rate_limiter = RateLimiter(calls_per_minute=10, max_workers=3)
        with concurrent.futures.ThreadPoolExecutor() as executor: # max_workers=2
            futures = {executor.submit(rate_limiter(summarize), c[2].target, max_word, chat_model, model_info): (c[2].target, c[3], c[0],c[1]) for c in coord}
            index = 0
            for future in concurrent.futures.as_completed(futures):
                url, title, col, row = futures[future]
                try:
                    my_bar.progress(int((index+1)/len(coord)*100))
                    result, cost = future.result()
                    cell = sheet.cell(column=col+2, row=row)
                    cell.value = result
                    st.write(f'[{title}]({url})')
                    st.write(result)
                    costs.append(cost)
                    index += 1
                except Exception as exc:
                    print(f'{url} generated an exception: {exc}')

        
        my_bar.progress(100)
        total = sum_cost(costs)
        logger.log(model, uploaded_file.name, total.input_token, total.input_cost, total.output_token, total.output_cost, total.total_cost)
        
        
        st.write('### Completed !')
        ttk = total.input_token + total.output_token
        tic = '{:.5f}'.format(total.input_cost)
        toc = '{:.5f}'.format(total.output_cost)
        tc  = '{:.5f}'.format(total.input_cost + total.output_cost)
        st.sidebar.write(f'''
        #### Cost
        |   |tokens|cost($)|
        |---|---|---|
        |input |{total.input_token} |{tic} |
        |output|{total.output_token}|{toc} |
        |total |{ttk}               |{tc}  |
        ''')
        
        outputfile = f'output/{uploaded_file.name}'
        book.save(outputfile)

        with open(outputfile, 'rb') as f:
            st.download_button(label='Download XLSX', 
                               data=f, 
                               file_name=uploaded_file.name, 
                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')  # Defaults to 'text/plain'
        
        
